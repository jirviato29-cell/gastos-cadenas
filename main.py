from flask import Flask, render_template, request, jsonify, redirect
import os
from datetime import date, datetime, timedelta
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2 import pool
import openpyxl
import xlrd
from io import BytesIO

app = Flask(__name__)
app.config['SECRET_KEY'] = 'gastosCadenas2026'

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres:password@localhost:5432/gastos"
)
# Railway entrega URLs con prefijo "postgres://" — psycopg2 requiere "postgresql://"
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

try:
    connection_pool = pool.ThreadedConnectionPool(2, 10, DATABASE_URL)
except Exception as e:
    print(f"Error conectando a DB: {e}")
    connection_pool = None


def get_conn():
    if connection_pool is None:
        raise RuntimeError("Sin conexión a la base de datos")
    conn = connection_pool.getconn()
    conn.cursor_factory = RealDictCursor
    return conn


def release_conn(conn):
    connection_pool.putconn(conn)


# ── Cálculos ─────────────────────────────────────────────────────────────────

def dias_vacaciones_ley(anos):
    if anos < 1:  return 0
    if anos < 2:  return 12
    if anos < 3:  return 14
    if anos < 4:  return 16
    if anos < 5:  return 18
    if anos < 10: return 20
    return 22


def calcular_gasto_promotor(p):
    fi = p.get('fecha_ingreso')
    if not fi:
        return {
            'sueldo_semanal': 0, 'aguinaldo': 0, 'vacaciones': 0,
            'prima_vacacional': 0, 'seguro': 0, 'isn': 0,
            'impuestos': 0, 'gastos_indirectos': 0,
            'fondo_contingencia': 0, 'total': 0,
            'anos': 0, 'dias_vac': 0,
        }
    fi = datetime.strptime(str(fi)[:10], '%Y-%m-%d').date()
    ss  = float(p['sueldo'] or 0)        # ya es semanal
    sd  = round(ss / 7, 2)               # sueldo diario = sueldo semanal ÷ 7
    anos      = max(0, (date.today() - fi).days / 365.25)
    aguinaldo = round((sd * 15) / 52, 2)
    if anos < 1:
        vacaciones, prima_vacacional, dias_vac = 0.0, 0.0, 0
    else:
        dias_vac         = dias_vacaciones_ley(int(anos))
        vacaciones       = round((dias_vac * sd) / 52, 2)
        prima_vacacional = round(vacaciones * 0.25, 2)
    seguro           = 696.42 if p.get('tiene_seguro') else 0.0
    isn              = 61.54
    impuestos        = 417.05
    gastos_indirectos   = 274.0
    fondo_contingencia  = 27.0
    total = round(ss + seguro + isn + impuestos + gastos_indirectos +
                  fondo_contingencia + aguinaldo + vacaciones + prima_vacacional, 2)
    return dict(sueldo_semanal=ss, seguro=seguro, isn=isn, impuestos=impuestos,
                gastos_indirectos=gastos_indirectos, fondo_contingencia=fondo_contingencia,
                aguinaldo=aguinaldo, vacaciones=vacaciones,
                prima_vacacional=prima_vacacional, total=total,
                anos=round(anos, 1), dias_vac=dias_vac)


# ── Init DB ───────────────────────────────────────────────────────────────────

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT pg_advisory_lock(88888)")
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tiendas (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                cadena TEXT NOT NULL,
                clave_principal TEXT NOT NULL UNIQUE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS promotores (
                id SERIAL PRIMARY KEY,
                promotor_id TEXT,
                nombre TEXT NOT NULL,
                tienda_id INTEGER REFERENCES tiendas(id),
                sueldo FLOAT NOT NULL DEFAULT 0,
                comisiones FLOAT DEFAULT 0,
                fecha_ingreso DATE,
                tiene_seguro BOOLEAN DEFAULT FALSE,
                dias_vacaciones_tomados INTEGER DEFAULT 0
            )
        """)
        cur.execute("""
            ALTER TABLE promotores ADD COLUMN IF NOT EXISTS promotor_id TEXT
        """)
        cur.execute("""
            ALTER TABLE promotores ADD COLUMN IF NOT EXISTS comisiones FLOAT DEFAULT 0
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS subclaves_telcel (
                id SERIAL PRIMARY KEY,
                tienda_id INTEGER REFERENCES tiendas(id) ON DELETE CASCADE,
                subclave TEXT NOT NULL,
                UNIQUE(tienda_id, subclave)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS semanas (
                id SERIAL PRIMARY KEY,
                fecha_inicio DATE NOT NULL,
                fecha_fin DATE NOT NULL,
                UNIQUE(fecha_inicio, fecha_fin)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS gastos_semana (
                id SERIAL PRIMARY KEY,
                semana_id INTEGER REFERENCES semanas(id) ON DELETE CASCADE,
                promotor_id INTEGER REFERENCES promotores(id) ON DELETE CASCADE,
                sueldo_semanal FLOAT DEFAULT 0,
                comisiones FLOAT DEFAULT 0,
                seguro FLOAT DEFAULT 0,
                isn FLOAT DEFAULT 0,
                impuestos FLOAT DEFAULT 0,
                gastos_indirectos FLOAT DEFAULT 0,
                fondo_contingencia FLOAT DEFAULT 0,
                aguinaldo FLOAT DEFAULT 0,
                vacaciones FLOAT DEFAULT 0,
                prima_vacacional FLOAT DEFAULT 0,
                total FLOAT DEFAULT 0,
                UNIQUE(semana_id, promotor_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS pagos_telcel (
                id SERIAL PRIMARY KEY,
                semana_id INTEGER REFERENCES semanas(id) ON DELETE CASCADE,
                subclave TEXT NOT NULL,
                producto TEXT,
                comision FLOAT DEFAULT 0
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS comisiones_extra (
                id SERIAL PRIMARY KEY,
                semana_id INTEGER REFERENCES semanas(id) ON DELETE CASCADE,
                tipo TEXT NOT NULL,
                monto FLOAT NOT NULL DEFAULT 0,
                notas TEXT
            )
        """)
        conn.commit()
        print("DB gastos-cadenas OK")
    finally:
        cur.execute("SELECT pg_advisory_unlock(88888)")
        cur.close()
        release_conn(conn)


# ── TIENDAS ───────────────────────────────────────────────────────────────────

@app.route('/tiendas')
def tiendas():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.*,
               COALESCE(json_agg(s.subclave ORDER BY s.id)
                        FILTER (WHERE s.id IS NOT NULL), '[]') AS subclaves
        FROM tiendas t
        LEFT JOIN subclaves_telcel s ON s.tienda_id = t.id
        GROUP BY t.id ORDER BY t.cadena, t.nombre
    """)
    rows = cur.fetchall()
    cur.close()
    release_conn(conn)
    return render_template('tiendas.html', tiendas=rows)


@app.route('/api/tiendas', methods=['POST'])
def add_tienda():
    d = request.json
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO tiendas (nombre, cadena, clave_principal) VALUES (%s,%s,%s) RETURNING id",
            (d['nombre'].strip(), d['cadena'].strip(), d['clave_principal'].strip()))
        tid = cur.fetchone()['id']
        for sub in d.get('subclaves', []):
            if sub.strip():
                cur.execute(
                    "INSERT INTO subclaves_telcel (tienda_id, subclave) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (tid, sub.strip()))
        conn.commit()
        return jsonify({'ok': True, 'id': tid})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/tiendas/<int:tid>', methods=['PUT'])
def edit_tienda(tid):
    d = request.json
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE tiendas SET nombre=%s, cadena=%s, clave_principal=%s WHERE id=%s",
            (d['nombre'].strip(), d['cadena'].strip(), d['clave_principal'].strip(), tid))
        cur.execute("DELETE FROM subclaves_telcel WHERE tienda_id=%s", (tid,))
        for sub in d.get('subclaves', []):
            if sub.strip():
                cur.execute(
                    "INSERT INTO subclaves_telcel (tienda_id, subclave) VALUES (%s,%s)",
                    (tid, sub.strip()))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/tiendas/<int:tid>', methods=['DELETE'])
def del_tienda(tid):
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM tiendas WHERE id=%s", (tid,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


# ── PROMOTORES ────────────────────────────────────────────────────────────────

@app.route('/promotores')
def promotores():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.*, t.nombre AS tienda_nombre, t.cadena
        FROM promotores p LEFT JOIN tiendas t ON t.id = p.tienda_id
        ORDER BY t.cadena NULLS LAST, t.nombre NULLS LAST, p.nombre
    """)
    proms = cur.fetchall()
    cur.execute("SELECT id, nombre, cadena FROM tiendas ORDER BY cadena, nombre")
    tiendas = cur.fetchall()
    cur.close()
    release_conn(conn)

    # Enriquecer con cálculos
    resultado = []
    for p in proms:
        p = dict(p)
        g = calcular_gasto_promotor(p)
        p.update(g)
        if p['fecha_ingreso']:
            p['fecha_ingreso_str'] = p['fecha_ingreso'].strftime('%Y-%m-%d')
        resultado.append(p)

    return render_template('promotores.html', promotores=resultado, tiendas=tiendas)


@app.route('/api/promotores', methods=['POST'])
def add_promotor():
    d = request.json
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO promotores (nombre, tienda_id, sueldo, fecha_ingreso,
                                    tiene_seguro, dias_vacaciones_tomados)
            VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
        """, (d['nombre'].strip(), d.get('tienda_id') or None,
              float(d['sueldo']), d['fecha_ingreso'],
              bool(d.get('tiene_seguro', False)),
              int(d.get('dias_vacaciones_tomados', 0))))
        pid = cur.fetchone()['id']
        conn.commit()
        return jsonify({'ok': True, 'id': pid})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/promotores/<int:pid>', methods=['PUT'])
def edit_promotor(pid):
    d = request.json
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE promotores SET nombre=%s, tienda_id=%s, sueldo=%s,
                fecha_ingreso=%s, tiene_seguro=%s, dias_vacaciones_tomados=%s
            WHERE id=%s
        """, (d['nombre'].strip(), d.get('tienda_id') or None,
              float(d['sueldo']), d['fecha_ingreso'],
              bool(d.get('tiene_seguro', False)),
              int(d.get('dias_vacaciones_tomados', 0)), pid))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/promotores/<int:pid>', methods=['DELETE'])
def del_promotor(pid):
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM promotores WHERE id=%s", (pid,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/promotores/importar', methods=['POST'])
def importar_promotores():
    f = request.files.get('archivo')
    if not f:
        return jsonify({'ok': False, 'error': 'No se recibió archivo'}), 400
    try:
        filename = f.filename.lower()
        raw = f.read()
        all_rows = []
        if filename.endswith('.xls'):
            wb_xls = xlrd.open_workbook(file_contents=raw)
            ws_xls = wb_xls.sheet_by_index(0)
            for i in range(1, ws_xls.nrows):
                all_rows.append([ws_xls.cell_value(i, c) if c < ws_xls.ncols else None for c in range(8)])
        else:
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
            ws = wb.active
            for row in list(ws.iter_rows(min_row=2, values_only=True)):
                all_rows.append(list(row[:8]))

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, nombre FROM tiendas")
        tiendas_all = cur.fetchall()
        tiendas_idx = {r['nombre'].strip().lower(): r['id'] for r in tiendas_all}

        def buscar_tienda(nombre_excel):
            clave = nombre_excel.strip().lower()
            if clave in tiendas_idx:
                return tiendas_idx[clave]
            for nom, tid in tiendas_idx.items():
                if clave and (clave in nom or nom.endswith(clave)):
                    return tid
            return None

        insertados = actualizados = errores = 0
        errores_detalle = []

        for row in all_rows:
            # Formato SUELDOS: B=Cadena, C=Tienda, D=ID corto, E=CLAVE (promotor_id), F=Nombre, K=Sueldo
            tienda_nombre = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            promotor_id   = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            nombre        = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            sueldo_raw    = row[10] if len(row) > 10 else None

            if not promotor_id or not nombre or promotor_id in ('CLAVE', 'ID', 'None', ''):
                continue

            tienda_id = buscar_tienda(tienda_nombre)
            if not tienda_id and tienda_nombre:
                errores_detalle.append(f'{promotor_id}: tienda "{tienda_nombre}" no encontrada')

            try:
                sueldo = float(sueldo_raw or 0)
            except (TypeError, ValueError):
                sueldo = 0.0

            cur.execute("SELECT id FROM promotores WHERE promotor_id=%s LIMIT 1", (promotor_id,))
            existing = cur.fetchone()

            try:
                if existing:
                    cur.execute("""
                        UPDATE promotores SET nombre=%s, tienda_id=%s, sueldo=%s, tiene_seguro=FALSE
                        WHERE id=%s
                    """, (nombre, tienda_id, sueldo, existing['id']))
                    actualizados += 1
                else:
                    cur.execute("""
                        INSERT INTO promotores (promotor_id, nombre, tienda_id, sueldo,
                            fecha_ingreso, tiene_seguro, dias_vacaciones_tomados)
                        VALUES (%s,%s,%s,%s,NULL,FALSE,0)
                    """, (promotor_id, nombre, tienda_id, sueldo))
                    insertados += 1
            except Exception as e:
                errores += 1
                errores_detalle.append(f'{promotor_id}: {e}')

        conn.commit()
        cur.close()
        release_conn(conn)
        return jsonify({'ok': True, 'insertados': insertados, 'actualizados': actualizados,
                        'errores': errores, 'errores_detalle': errores_detalle[:10]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── GASTOS ────────────────────────────────────────────────────────────────────

@app.route('/gastos')
def gastos():
    return render_template('gastos.html')


@app.route('/api/semanas', methods=['POST'])
def add_semana():
    d = request.json
    fi = datetime.strptime(d['fecha_inicio'], '%Y-%m-%d').date()
    lunes = fi - timedelta(days=fi.weekday())
    domingo = lunes + timedelta(days=6)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO semanas (fecha_inicio, fecha_fin) VALUES (%s,%s) RETURNING id",
            (lunes.isoformat(), domingo.isoformat()))
        sid = cur.fetchone()['id']
        conn.commit()
        return jsonify({'ok': True, 'id': sid})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/semanas', methods=['GET'])
def list_semanas():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC")
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        r['fecha_inicio'] = str(r['fecha_inicio'])
        r['fecha_fin'] = str(r['fecha_fin'])
    cur.close()
    release_conn(conn)
    return jsonify(rows)


@app.route('/api/gastos/generar', methods=['POST'])
def generar_gastos():
    semana_id = int(request.json['semana_id'])
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM promotores")
        proms = cur.fetchall()

        # Comisiones Telcel totales por tienda para esta semana
        cur.execute("""
            SELECT st.tienda_id, SUM(pt.comision) AS total
            FROM pagos_telcel pt
            JOIN subclaves_telcel st ON st.subclave = pt.subclave
            WHERE pt.semana_id = %s
            GROUP BY st.tienda_id
        """, (semana_id,))
        com_tienda = {r['tienda_id']: float(r['total'] or 0) for r in cur.fetchall()}

        for p in proms:
            g = calcular_gasto_promotor(p)
            # impuestos, gastos_indirectos y fondo_contingencia son fijos por tienda;
            # se guardan en 0 por promotor y se inyectan como constante en get_gastos()
            total_promotor = round(
                g['sueldo_semanal'] + g['seguro'] + g['isn'] +
                g['aguinaldo'] + g['vacaciones'] + g['prima_vacacional'], 2
            )
            com = round(float(p.get('comisiones') or 0), 2)
            total_final = round(total_promotor + com, 2)
            cur.execute("""
                INSERT INTO gastos_semana
                    (semana_id, promotor_id, sueldo_semanal, comisiones, seguro, isn,
                     impuestos, gastos_indirectos, fondo_contingencia,
                     aguinaldo, vacaciones, prima_vacacional, total)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (semana_id, promotor_id) DO UPDATE SET
                    sueldo_semanal=EXCLUDED.sueldo_semanal,
                    comisiones=EXCLUDED.comisiones,
                    seguro=EXCLUDED.seguro, isn=EXCLUDED.isn,
                    impuestos=EXCLUDED.impuestos,
                    gastos_indirectos=EXCLUDED.gastos_indirectos,
                    fondo_contingencia=EXCLUDED.fondo_contingencia,
                    aguinaldo=EXCLUDED.aguinaldo, vacaciones=EXCLUDED.vacaciones,
                    prima_vacacional=EXCLUDED.prima_vacacional, total=EXCLUDED.total
            """, (semana_id, p['id'], g['sueldo_semanal'], com, g['seguro'],
                  g['isn'], 0, 0, 0,
                  g['aguinaldo'], g['vacaciones'], g['prima_vacacional'], total_final))

        conn.commit()
        return jsonify({'ok': True, 'promotores': len(proms)})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        cur.close()
        release_conn(conn)


@app.route('/api/gastos/<int:semana_id>')
def get_gastos(semana_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT t.cadena, t.nombre AS tienda,
               COUNT(gs.id)             AS promotores,
               SUM(gs.sueldo_semanal)   AS sueldos,
               SUM(gs.comisiones)       AS comisiones,
               SUM(gs.seguro)           AS seguro,
               SUM(gs.aguinaldo)        AS aguinaldo,
               SUM(gs.vacaciones)       AS vacaciones,
               SUM(gs.prima_vacacional) AS prima_vacacional,
               SUM(gs.total)            AS total
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        JOIN tiendas t ON t.id = p.tienda_id
        WHERE gs.semana_id = %s
        GROUP BY t.id, t.cadena, t.nombre
        ORDER BY t.cadena, t.nombre
    """, (semana_id,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    release_conn(conn)
    for r in rows:
        for k, v in r.items():
            if v is not None and k not in ('cadena', 'tienda'):
                try:
                    r[k] = round(float(v), 2)
                except Exception:
                    pass
        n = int(r['promotores'])
        r['isn']                = round(61.54 * n, 2)
        r['impuestos']          = 417.05
        r['gastos_indirectos']  = 274.00
        r['fondo_contingencia'] = 27.00
        r['total']              = round(float(r['total'] or 0) + 417.05 + 274.00 + 27.00, 2)
    return jsonify(rows)


@app.route('/api/gastos/<int:semana_id>/detalle')
def get_gastos_detalle(semana_id):
    """Gastos por promotor individual para las pestañas de la UI."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT gs.sueldo_semanal, gs.comisiones, gs.seguro, gs.isn,
               gs.impuestos, gs.gastos_indirectos, gs.fondo_contingencia,
               gs.aguinaldo, gs.vacaciones, gs.prima_vacacional, gs.total,
               p.nombre AS promotor, p.promotor_id,
               t.nombre AS tienda,
               t.cadena
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        JOIN tiendas t ON t.id = p.tienda_id
        WHERE gs.semana_id = %s
        ORDER BY t.cadena, t.nombre, p.promotor_id
    """, (semana_id,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    release_conn(conn)
    for r in rows:
        for k, v in r.items():
            if k not in ('promotor', 'tienda', 'cadena') and v is not None:
                try:
                    r[k] = round(float(v), 2)
                except Exception:
                    pass
    return jsonify(rows)


# ── TELCEL ────────────────────────────────────────────────────────────────────

@app.route('/telcel')
def telcel():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC")
    semanas = cur.fetchall()
    cur.close()
    release_conn(conn)
    return render_template('telcel.html', semanas=semanas)


@app.route('/api/telcel/upload', methods=['POST'])
def upload_telcel():
    semana_id = request.form.get('semana_id')
    if not semana_id:
        return jsonify({'ok': False, 'error': 'Selecciona una semana'}), 400
    f = request.files.get('archivo')
    if not f:
        return jsonify({'ok': False, 'error': 'No se recibió archivo'}), 400

    try:
        filename = f.filename.lower()
        raw = f.read()

        # Columnas fijas: A=Subclave, B=Concepto, C=Comisión
        all_rows = []
        if filename.endswith('.xls'):
            wb_xls = xlrd.open_workbook(file_contents=raw)
            ws_xls = wb_xls.sheet_by_index(0)
            for i in range(ws_xls.nrows):
                all_rows.append([ws_xls.cell_value(i, c) if c < ws_xls.ncols else None
                                  for c in range(3)])
        else:
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row[:3]))

        if not all_rows:
            return jsonify({'ok': False, 'error': 'El archivo está vacío'}), 400

        # Detectar si la primera fila es encabezado (col C no numérica)
        try:
            float(all_rows[0][2] or 0)
            start = 0
        except (TypeError, ValueError):
            start = 1

        rows_data = []
        for row in all_rows[start:]:
            if len(row) < 3:
                continue
            sub  = str(row[0]).strip() if row[0] is not None else ''
            prod = str(row[1]).strip() if row[1] is not None else ''
            try:
                com = float(row[2] or 0)
            except (TypeError, ValueError):
                com = 0.0
            if sub and sub.lower() not in ('none', 'nan', ''):
                rows_data.append((sub, prod, com))

        if not rows_data:
            return jsonify({'ok': False, 'error': 'El archivo no contiene datos válidos'}), 400

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM pagos_telcel WHERE semana_id=%s", (semana_id,))
        cur.executemany(
            "INSERT INTO pagos_telcel (semana_id, subclave, producto, comision) VALUES (%s,%s,%s,%s)",
            [(semana_id, r[0], r[1], r[2]) for r in rows_data])
        conn.commit()

        # Conceptos únicos en orden de aparición
        conceptos = []
        seen_c = set()
        for r in rows_data:
            c = r[1]
            if c and c not in seen_c:
                seen_c.add(c)
                conceptos.append(c)

        cur.execute("""
            SELECT pt.subclave,
                   COALESCE(t.nombre,'Sin tienda asignada') AS tienda,
                   COALESCE(t.cadena,'') AS cadena,
                   pt.producto,
                   SUM(pt.comision) AS subtotal
            FROM pagos_telcel pt
            LEFT JOIN subclaves_telcel st ON st.subclave = pt.subclave
            LEFT JOIN tiendas t ON t.id = st.tienda_id
            WHERE pt.semana_id = %s
            GROUP BY pt.subclave, t.nombre, t.cadena, pt.producto
            ORDER BY pt.subclave
        """, (semana_id,))
        raw_rows = cur.fetchall()
        cur.close()
        release_conn(conn)

        # Agrupar por subclave con un dict por concepto
        agrupado = {}
        orden_sub = []
        for row in raw_rows:
            sub = row['subclave']
            if sub not in agrupado:
                agrupado[sub] = {
                    'subclave': sub,
                    'tienda': row['tienda'],
                    'cadena': row['cadena'],
                    'conceptos': {}
                }
                orden_sub.append(sub)
            prod = row['producto'] or ''
            agrupado[sub]['conceptos'][prod] = round(float(row['subtotal'] or 0), 2)

        resumen = []
        for sub in orden_sub:
            r = agrupado[sub]
            total = sum(r['conceptos'].values())
            resumen.append({
                'subclave': r['subclave'],
                'tienda':   r['tienda'],
                'cadena':   r['cadena'],
                'conceptos': r['conceptos'],
                'total':    round(total, 2)
            })

        return jsonify({'ok': True, 'filas': len(rows_data), 'conceptos': conceptos, 'resumen': resumen})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── COMISIONES ────────────────────────────────────────────────────────────────

@app.route('/comisiones')
def comisiones():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC")
    semanas = cur.fetchall()
    cur.close()
    release_conn(conn)
    return render_template('comisiones.html', semanas=semanas)


CONCEPTOS_FIJOS = ['AMIGO KIT', 'CBPC', 'CHIP EXPRESS', 'GARANTIZADA']

@app.route('/api/comisiones/<int:semana_id>')
def get_comisiones(semana_id):
    conn = get_conn()
    cur = conn.cursor()
    # Agrupar por tienda (suma clave principal + todas sus subclaves)
    cur.execute("""
        SELECT COALESCE(t.cadena,'Sin cadena')        AS cadena,
               COALESCE(t.nombre,'Sin tienda asignada') AS tienda,
               COALESCE(t.clave_principal,'')          AS clave_principal,
               pt.producto,
               SUM(pt.comision)                        AS subtotal
        FROM pagos_telcel pt
        LEFT JOIN subclaves_telcel st ON st.subclave = pt.subclave
        LEFT JOIN tiendas t ON t.id = st.tienda_id
        WHERE pt.semana_id = %s
        GROUP BY t.cadena, t.nombre, t.clave_principal, pt.producto
        ORDER BY t.cadena, t.nombre
    """, (semana_id,))
    raw_rows = cur.fetchall()
    cur.close()
    release_conn(conn)

    # Agrupar en Python por (cadena, tienda)
    agrupado = {}
    orden = []
    for row in raw_rows:
        key = (row['cadena'], row['tienda'])
        if key not in agrupado:
            agrupado[key] = {'cadena': row['cadena'], 'tienda': row['tienda'],
                             'clave_principal': row['clave_principal'], 'conceptos': {}}
            orden.append(key)
        prod = (row['producto'] or '').strip().upper()
        agrupado[key]['conceptos'][prod] = round(float(row['subtotal'] or 0), 2)

    rows = []
    for key in orden:
        r = agrupado[key]
        total = round(sum(r['conceptos'].values()), 2)
        rows.append({'cadena': r['cadena'], 'tienda': r['tienda'],
                     'clave_principal': r['clave_principal'],
                     'conceptos': r['conceptos'], 'total': total})

    # Tiendas sin comisiones esa semana
    conn2 = get_conn()
    cur2 = conn2.cursor()
    cur2.execute("""
        SELECT DISTINCT t.id FROM tiendas t
        JOIN subclaves_telcel st ON st.tienda_id = t.id
        JOIN pagos_telcel pt ON pt.subclave = st.subclave
        WHERE pt.semana_id = %s
    """, (semana_id,))
    ids_con = {r['id'] for r in cur2.fetchall()}

    cur2.execute("SELECT id, cadena, nombre, clave_principal FROM tiendas ORDER BY cadena, nombre")
    todas = cur2.fetchall()
    cur2.close()
    release_conn(conn2)

    sin_comisiones = [
        {'cadena': t['cadena'], 'tienda': t['nombre'], 'clave_principal': t['clave_principal'] or ''}
        for t in todas if t['id'] not in ids_con
    ]

    return jsonify({'conceptos': CONCEPTOS_FIJOS, 'rows': rows, 'sin_comisiones': sin_comisiones})


# ── RESUMEN ───────────────────────────────────────────────────────────────────

@app.route('/resumen')
def resumen():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC")
    semanas = cur.fetchall()
    cur.close()
    release_conn(conn)
    return render_template('resumen.html', semanas=semanas)


@app.route('/api/resumen/<int:semana_id>')
def get_resumen(semana_id):
    conn = get_conn()
    cur = conn.cursor()

    # Gastos por tienda
    cur.execute("""
        SELECT t.id, t.cadena, t.nombre AS tienda,
               SUM(gs.total) AS gastos, COUNT(gs.id) AS promotores
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        JOIN tiendas t ON t.id = p.tienda_id
        WHERE gs.semana_id = %s
        GROUP BY t.id, t.cadena, t.nombre
    """, (semana_id,))
    result = {}
    for r in cur.fetchall():
        result[r['id']] = {
            'id': r['id'], 'cadena': r['cadena'], 'tienda': r['tienda'],
            'gastos': round(float(r['gastos'] or 0), 2),
            'promotores': int(r['promotores']), 'ingresos': 0.0
        }

    # Ingresos Telcel por tienda (via subclaves)
    cur.execute("""
        SELECT st.tienda_id, t.cadena, t.nombre AS tienda, SUM(pt.comision) AS ingresos
        FROM pagos_telcel pt
        JOIN subclaves_telcel st ON st.subclave = pt.subclave
        JOIN tiendas t ON t.id = st.tienda_id
        WHERE pt.semana_id = %s
        GROUP BY st.tienda_id, t.cadena, t.nombre
    """, (semana_id,))
    for r in cur.fetchall():
        tid = r['tienda_id']
        ing = round(float(r['ingresos'] or 0), 2)
        if tid in result:
            result[tid]['ingresos'] = ing
        else:
            result[tid] = {
                'id': tid, 'cadena': r['cadena'], 'tienda': r['tienda'],
                'gastos': 0.0, 'promotores': 0, 'ingresos': ing
            }

    rows = sorted(result.values(), key=lambda x: (x['cadena'], x['tienda']))
    for r in rows:
        r['utilidad'] = round(r['ingresos'] - r['gastos'], 2)

    cur.close()
    release_conn(conn)
    return jsonify(rows)


# ── BALANCE ───────────────────────────────────────────────────────────────────

@app.route('/balance')
def balance():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC")
    semanas = [dict(r) for r in cur.fetchall()]
    for s in semanas:
        s['fecha_inicio'] = str(s['fecha_inicio']); s['fecha_fin'] = str(s['fecha_fin'])
    cur.close(); release_conn(conn)
    return render_template('balance.html', semanas=semanas)


@app.route('/api/balance/<int:semana_id>')
def get_balance(semana_id):
    conn = get_conn(); cur = conn.cursor()

    cur.execute("""
        SELECT UPPER(TRIM(producto)) AS producto, SUM(comision) AS total, COUNT(*) AS lineas
        FROM pagos_telcel WHERE semana_id = %s
        GROUP BY UPPER(TRIM(producto))
    """, (semana_id,))
    telcel_rows = cur.fetchall()

    cur.execute("""
        SELECT tipo, SUM(monto) AS total
        FROM comisiones_extra WHERE semana_id = %s
        GROUP BY tipo
    """, (semana_id,))
    extra_rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(gs.id)                                        AS n_prom,
               COUNT(DISTINCT p.tienda_id)                         AS n_tiend,
               SUM(gs.sueldo_semanal)                              AS sueldos,
               SUM(gs.comisiones)                                  AS comisiones,
               SUM(CASE WHEN gs.comisiones > 0 THEN 1 ELSE 0 END)  AS n_com,
               SUM(gs.seguro)                                      AS seguro,
               SUM(CASE WHEN gs.seguro > 0 THEN 1 ELSE 0 END)     AS n_seg,
               SUM(gs.aguinaldo)                                   AS aguinaldo,
               SUM(gs.vacaciones)                                  AS vacaciones,
               SUM(CASE WHEN gs.vacaciones > 0 THEN 1 ELSE 0 END) AS n_vac,
               SUM(gs.prima_vacacional)                            AS prima_vacacional,
               SUM(CASE WHEN gs.prima_vacacional > 0 THEN 1 ELSE 0 END) AS n_pv
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        WHERE gs.semana_id = %s
    """, (semana_id,))
    g = dict(cur.fetchone())
    cur.close(); release_conn(conn)

    if not g or not g['n_prom']:
        return jsonify({'ok': False, 'error': 'Sin gastos generados para esta semana'})

    np  = int(g['n_prom']  or 0)
    nt  = int(g['n_tiend'] or 0)

    telcel = {}
    telcel_lineas = {}
    for r in telcel_rows:
        prod = (r['producto'] or '').strip()
        if prod:
            telcel[prod] = round(float(r['total'] or 0), 2)
            telcel_lineas[prod] = int(r['lineas'] or 0)

    extra = {r['tipo']: round(float(r['total'] or 0), 2) for r in extra_rows}

    total_ingresos = round(sum(telcel.values()) + sum(extra.values()), 2)

    sueldos = round(float(g['sueldos'] or 0), 2)
    coms    = round(float(g['comisiones'] or 0), 2)
    seguro  = round(float(g['seguro'] or 0), 2)
    isn     = round(61.54  * np, 2)
    imp     = round(417.05 * nt, 2)
    gi      = round(274.00 * nt, 2)
    fc      = round(27.00  * np, 2)
    ag      = round(float(g['aguinaldo'] or 0), 2)
    vac     = round(float(g['vacaciones'] or 0), 2)
    pv      = round(float(g['prima_vacacional'] or 0), 2)

    total_gastos = round(sueldos + coms + seguro + isn + imp + gi + fc + ag + vac + pv, 2)
    utilidad     = round(total_ingresos - total_gastos, 2)
    margen       = round(utilidad / total_ingresos * 100, 1) if total_ingresos else 0.0

    return jsonify({
        'ok': True,
        'ingresos': {'telcel': telcel, 'telcel_lineas': telcel_lineas, 'extra': extra, 'total': total_ingresos},
        'gastos': {
            'sueldos': sueldos,    'n_prom': np,
            'comisiones': coms,    'n_com': int(g['n_com'] or 0),
            'seguro': seguro,      'n_seg': int(g['n_seg'] or 0),
            'isn': isn,
            'impuestos': imp,      'n_tiend': nt,
            'gastos_indirectos': gi,
            'fondo_contingencia': fc,
            'aguinaldo': ag,
            'vacaciones': vac,     'n_vac': int(g['n_vac'] or 0),
            'prima_vacacional': pv,'n_pv': int(g['n_pv'] or 0),
            'total': total_gastos,
        },
        'utilidad': utilidad,
        'margen': margen,
    })


# ── COMISIONES EXTRA ─────────────────────────────────────────────────────────

TIPOS_COMISION_EXTRA = ['2% AP', 'Volumen Garantizado', 'Amigo Kit', 'Cadena Comercial']

@app.route('/comisiones-extra')
def comisiones_extra():
    conn = get_conn(); cur = conn.cursor()
    cur.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC")
    semanas = [dict(r) for r in cur.fetchall()]
    for s in semanas:
        s['fecha_inicio'] = str(s['fecha_inicio']); s['fecha_fin'] = str(s['fecha_fin'])
    cur.close(); release_conn(conn)
    return render_template('comisiones_extra.html', semanas=semanas,
                           tipos=TIPOS_COMISION_EXTRA)


@app.route('/api/comisiones-extra/<int:semana_id>')
def get_comisiones_extra(semana_id):
    conn = get_conn(); cur = conn.cursor()
    cur.execute("""
        SELECT id, tipo, monto, notas
        FROM comisiones_extra
        WHERE semana_id = %s
        ORDER BY tipo
    """, (semana_id,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); release_conn(conn)
    for r in rows:
        r['monto'] = round(float(r['monto'] or 0), 2)
    return jsonify(rows)


@app.route('/api/comisiones-extra', methods=['POST'])
def add_comision_extra():
    d = request.json
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO comisiones_extra (semana_id, tipo, monto, notas)
            VALUES (%s,%s,%s,%s) RETURNING id
        """, (int(d['semana_id']), d['tipo'].strip(),
              float(d['monto']), (d.get('notas') or '').strip() or None))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'ok': True, 'id': new_id})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close(); release_conn(conn)


@app.route('/api/comisiones-extra/<int:cid>', methods=['PUT'])
def edit_comision_extra(cid):
    d = request.json
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE comisiones_extra SET tipo=%s, monto=%s, notas=%s
            WHERE id=%s
        """, (d['tipo'].strip(), float(d['monto']),
              (d.get('notas') or '').strip() or None, cid))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close(); release_conn(conn)


@app.route('/api/comisiones-extra/<int:cid>', methods=['DELETE'])
def del_comision_extra(cid):
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM comisiones_extra WHERE id=%s", (cid,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        cur.close(); release_conn(conn)


# ── Root ──────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect('/gastos')


try:
    init_db()
except Exception as e:
    print(f"init_db: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
