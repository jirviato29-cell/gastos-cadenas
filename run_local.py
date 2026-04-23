"""
run_local.py — versión SQLite para pruebas locales.
Usa la misma carpeta templates/ que main.py.
Ejecutar: python run_local.py
"""
import sqlite3
import os
from flask import Flask, render_template, request, jsonify, redirect
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
import xlrd

app = Flask(__name__, template_folder='templates')
app.config['SECRET_KEY'] = 'local-test-2026'

DB_PATH = os.path.join(os.path.dirname(__file__), 'local_gastos.db')


# ── SQLite helpers ────────────────────────────────────────────────────────────

def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db

def ql(rows): return [dict(r) for r in rows]
def q1(row):  return dict(row) if row else None


# ── Cálculos (idénticos a main.py) ───────────────────────────────────────────

def dias_vacaciones_ley(anos):
    if anos < 1:  return 0
    if anos < 2:  return 12
    if anos < 3:  return 14
    if anos < 4:  return 16
    if anos < 5:  return 18
    if anos < 10: return 20
    return 22


def calcular_gasto_promotor(p):
    fi = p.get('fecha_ingreso')
    if not fi:
        return {
            'sueldo_semanal': 0, 'aguinaldo': 0, 'vacaciones': 0,
            'prima_vacacional': 0, 'seguro': 0, 'isn': 0,
            'impuestos': 0, 'gastos_indirectos': 0,
            'fondo_contingencia': 0, 'total': 0,
            'anos': 0, 'dias_vac': 0,
        }
    fi = datetime.strptime(str(fi)[:10], '%Y-%m-%d').date()
    ss  = float(p['sueldo'] or 0)        # ya es semanal
    sd  = round(ss / 7, 2)               # sueldo diario = sueldo semanal ÷ 7
    anos = max(0, (date.today() - fi).days / 365.25)
    ag   = round((sd * 15) / 52, 2)
    if anos < 1:
        vac, pv, dias_vac = 0.0, 0.0, 0
    else:
        dias_vac = dias_vacaciones_ley(int(anos))
        vac = round((dias_vac * sd) / 52, 2)
        pv  = round(vac * 0.25, 2)
    seg = 696.42 if p.get('tiene_seguro') else 0.0
    isn, imp, gi, fc = 61.54, 417.05, 274.0, 27.0
    total = round(ss + seg + isn + imp + gi + fc + ag + vac + pv, 2)
    return dict(sueldo_semanal=ss, seguro=seg, isn=isn, impuestos=imp,
                gastos_indirectos=gi, fondo_contingencia=fc,
                aguinaldo=ag, vacaciones=vac, prima_vacacional=pv,
                total=total, anos=round(anos, 1), dias_vac=dias_vac)


# ── Init DB ───────────────────────────────────────────────────────────────────

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS tiendas (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre          TEXT NOT NULL,
            cadena          TEXT NOT NULL,
            clave_principal TEXT NOT NULL UNIQUE
        );
        CREATE TABLE IF NOT EXISTS promotores (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            promotor_id             TEXT,
            nombre                  TEXT NOT NULL,
            tienda_id               INTEGER REFERENCES tiendas(id),
            sueldo                  REAL NOT NULL DEFAULT 0,
            comisiones              REAL DEFAULT 0,
            fecha_ingreso           TEXT,
            tiene_seguro            INTEGER DEFAULT 0,
            dias_vacaciones_tomados INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS subclaves_telcel (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            tienda_id INTEGER REFERENCES tiendas(id) ON DELETE CASCADE,
            subclave  TEXT NOT NULL,
            UNIQUE(tienda_id, subclave)
        );
        CREATE TABLE IF NOT EXISTS semanas (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_inicio TEXT NOT NULL,
            fecha_fin    TEXT NOT NULL,
            UNIQUE(fecha_inicio, fecha_fin)
        );
        CREATE TABLE IF NOT EXISTS gastos_semana (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            semana_id          INTEGER REFERENCES semanas(id)   ON DELETE CASCADE,
            promotor_id        INTEGER REFERENCES promotores(id) ON DELETE CASCADE,
            sueldo_semanal     REAL DEFAULT 0,
            comisiones         REAL DEFAULT 0,
            seguro             REAL DEFAULT 0,
            isn                REAL DEFAULT 0,
            impuestos          REAL DEFAULT 0,
            gastos_indirectos  REAL DEFAULT 0,
            fondo_contingencia REAL DEFAULT 0,
            aguinaldo          REAL DEFAULT 0,
            vacaciones         REAL DEFAULT 0,
            prima_vacacional   REAL DEFAULT 0,
            total              REAL DEFAULT 0,
            UNIQUE(semana_id, promotor_id)
        );
        CREATE TABLE IF NOT EXISTS pagos_telcel (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            semana_id INTEGER REFERENCES semanas(id) ON DELETE CASCADE,
            subclave  TEXT NOT NULL,
            producto  TEXT,
            comision  REAL DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS comisiones_extra (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            semana_id INTEGER REFERENCES semanas(id) ON DELETE CASCADE,
            tipo      TEXT NOT NULL,
            monto     REAL NOT NULL DEFAULT 0,
            notas     TEXT
        );
    """)
    db.commit()
    try:
        db.execute("ALTER TABLE promotores ADD COLUMN promotor_id TEXT")
        db.commit()
    except Exception:
        pass
    try:
        db.execute("ALTER TABLE promotores ADD COLUMN fecha_ingreso TEXT")
        db.commit()
    except Exception:
        pass
    try:
        db.execute("ALTER TABLE promotores ADD COLUMN comisiones REAL DEFAULT 0")
        db.commit()
    except Exception:
        pass
    db.close()
    print("SQLite DB inicializada —", DB_PATH)


# ── TIENDAS ───────────────────────────────────────────────────────────────────

@app.route('/tiendas')
def tiendas():
    db = get_db()
    ts   = ql(db.execute("SELECT * FROM tiendas ORDER BY cadena, nombre").fetchall())
    subs = ql(db.execute("SELECT * FROM subclaves_telcel ORDER BY id").fetchall())
    db.close()
    smap = {}
    for s in subs:
        smap.setdefault(s['tienda_id'], []).append(s['subclave'])
    for t in ts:
        t['subclaves'] = smap.get(t['id'], [])
    return render_template('tiendas.html', tiendas=ts)


@app.route('/api/tiendas', methods=['POST'])
def add_tienda():
    d  = request.json
    db = get_db()
    try:
        cur = db.execute(
            "INSERT INTO tiendas (nombre, cadena, clave_principal) VALUES (?,?,?)",
            (d['nombre'].strip(), d['cadena'].strip(), d['clave_principal'].strip()))
        tid = cur.lastrowid
        for sub in d.get('subclaves', []):
            if sub.strip():
                db.execute("INSERT OR IGNORE INTO subclaves_telcel (tienda_id, subclave) VALUES (?,?)",
                           (tid, sub.strip()))
        db.commit()
        return jsonify({'ok': True, 'id': tid})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/tiendas/<int:tid>', methods=['PUT'])
def edit_tienda(tid):
    d  = request.json
    db = get_db()
    try:
        db.execute("UPDATE tiendas SET nombre=?, cadena=?, clave_principal=? WHERE id=?",
                   (d['nombre'].strip(), d['cadena'].strip(), d['clave_principal'].strip(), tid))
        db.execute("DELETE FROM subclaves_telcel WHERE tienda_id=?", (tid,))
        for sub in d.get('subclaves', []):
            if sub.strip():
                db.execute("INSERT INTO subclaves_telcel (tienda_id, subclave) VALUES (?,?)",
                           (tid, sub.strip()))
        db.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/tiendas/<int:tid>', methods=['DELETE'])
def del_tienda(tid):
    db = get_db()
    try:
        db.execute("DELETE FROM tiendas WHERE id=?", (tid,))
        db.commit(); return jsonify({'ok': True})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


# ── PROMOTORES ────────────────────────────────────────────────────────────────

@app.route('/promotores')
def promotores():
    db = get_db()
    proms   = ql(db.execute("""
        SELECT p.*, t.nombre AS tienda_nombre, t.cadena
        FROM promotores p LEFT JOIN tiendas t ON t.id = p.tienda_id
        ORDER BY t.cadena, t.nombre, p.nombre
    """).fetchall())
    tiendas = ql(db.execute(
        "SELECT id, nombre, cadena FROM tiendas ORDER BY cadena, nombre").fetchall())
    db.close()
    result = []
    for p in proms:
        p['tiene_seguro']     = bool(p['tiene_seguro'])
        p['fecha_ingreso_str'] = p['fecha_ingreso'] or ''
        g = calcular_gasto_promotor(p)
        p.update(g)
        result.append(p)
    return render_template('promotores.html', promotores=result, tiendas=tiendas)


@app.route('/api/promotores', methods=['POST'])
def add_promotor():
    d  = request.json
    db = get_db()
    try:
        cur = db.execute("""
            INSERT INTO promotores
                (nombre, tienda_id, sueldo, fecha_ingreso, tiene_seguro, dias_vacaciones_tomados)
            VALUES (?,?,?,?,?,?)
        """, (d['nombre'].strip(), d.get('tienda_id') or None,
              float(d['sueldo']), d['fecha_ingreso'],
              1 if d.get('tiene_seguro') else 0,
              int(d.get('dias_vacaciones_tomados', 0))))
        db.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/promotores/<int:pid>', methods=['PUT'])
def edit_promotor(pid):
    d  = request.json
    db = get_db()
    try:
        db.execute("""
            UPDATE promotores SET nombre=?, tienda_id=?, sueldo=?,
                fecha_ingreso=?, tiene_seguro=?, dias_vacaciones_tomados=?
            WHERE id=?
        """, (d['nombre'].strip(), d.get('tienda_id') or None,
              float(d['sueldo']), d['fecha_ingreso'],
              1 if d.get('tiene_seguro') else 0,
              int(d.get('dias_vacaciones_tomados', 0)), pid))
        db.commit(); return jsonify({'ok': True})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/promotores/<int:pid>', methods=['DELETE'])
def del_promotor(pid):
    db = get_db()
    try:
        db.execute("DELETE FROM promotores WHERE id=?", (pid,))
        db.commit(); return jsonify({'ok': True})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/promotores/importar', methods=['POST'])
def importar_promotores():
    f = request.files.get('archivo')
    if not f:
        return jsonify({'ok': False, 'error': 'No se recibió archivo'}), 400
    try:
        filename = f.filename.lower()
        raw = f.read()
        all_rows = []
        if filename.endswith('.xls'):
            wb_xls = xlrd.open_workbook(file_contents=raw)
            ws_xls = wb_xls.sheet_by_index(0)
            for i in range(1, ws_xls.nrows):  # saltar encabezado
                all_rows.append([ws_xls.cell_value(i, c) if c < ws_xls.ncols else None for c in range(8)])
        else:
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
            ws = wb.active
            for row in list(ws.iter_rows(min_row=2, values_only=True)):
                all_rows.append(list(row[:8]))

        db = get_db()
        # Índice de tiendas: nombre exacto y nombre parcial
        tiendas_all = ql(db.execute("SELECT id, nombre FROM tiendas").fetchall())
        tiendas_idx = {t['nombre'].strip().lower(): t['id'] for t in tiendas_all}

        def buscar_tienda(nombre_excel):
            clave = nombre_excel.strip().lower()
            if clave in tiendas_idx:
                return tiendas_idx[clave]
            # Coincidencia parcial: el nombre de la DB contiene el del Excel
            for nom, tid in tiendas_idx.items():
                if clave and (clave in nom or nom.endswith(clave)):
                    return tid
            return None

        insertados = actualizados = errores = 0
        errores_detalle = []

        for row in all_rows:
            # Formato SUELDOS: B=Cadena, C=Tienda, D=ID corto, E=CLAVE (promotor_id), F=Nombre, K=Sueldo
            tienda_nombre = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            promotor_id   = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            nombre        = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            sueldo_raw    = row[10] if len(row) > 10 else None

            # Saltar filas sin promotor_id válido (encabezados, filas vacías, subtotales)
            if not promotor_id or not nombre or promotor_id in ('CLAVE', 'ID', 'None', ''):
                continue

            tienda_id = buscar_tienda(tienda_nombre)
            if not tienda_id and tienda_nombre:
                errores_detalle.append(f'{promotor_id}: tienda "{tienda_nombre}" no encontrada')

            try:
                sueldo = float(sueldo_raw or 0)
            except (TypeError, ValueError):
                sueldo = 0.0

            existing = db.execute(
                "SELECT id FROM promotores WHERE promotor_id=? LIMIT 1", (promotor_id,)
            ).fetchone()

            try:
                if existing:
                    db.execute("""
                        UPDATE promotores SET nombre=?, tienda_id=?, sueldo=?, tiene_seguro=0
                        WHERE id=?
                    """, (nombre, tienda_id, sueldo, existing['id']))
                    actualizados += 1
                else:
                    db.execute("""
                        INSERT INTO promotores (promotor_id, nombre, tienda_id, sueldo,
                            fecha_ingreso, tiene_seguro, dias_vacaciones_tomados)
                        VALUES (?,?,?,?,NULL,0,0)
                    """, (promotor_id, nombre, tienda_id, sueldo))
                    insertados += 1
            except Exception as e:
                errores += 1
                errores_detalle.append(f'{promotor_id}: {e}')

        db.commit()
        db.close()
        return jsonify({'ok': True, 'insertados': insertados, 'actualizados': actualizados,
                        'errores': errores, 'errores_detalle': errores_detalle[:10]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── SEMANAS Y GASTOS ──────────────────────────────────────────────────────────

@app.route('/gastos')
def gastos():
    db      = get_db()
    semanas = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return render_template('gastos.html', semanas=semanas)


@app.route('/api/semanas', methods=['GET'])
def list_semanas():
    db   = get_db()
    rows = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return jsonify(rows)


@app.route('/api/semanas', methods=['POST'])
def add_semana():
    d  = request.json
    fi = datetime.strptime(d['fecha_inicio'], '%Y-%m-%d').date()
    lunes = fi - timedelta(days=fi.weekday())
    domingo = lunes + timedelta(days=6)
    print(f"Fecha recibida: {fi}, Lunes calculado: {lunes}")
    db = get_db()
    try:
        cur = db.execute("INSERT INTO semanas (fecha_inicio, fecha_fin) VALUES (?,?)",
                         (lunes.isoformat(), domingo.isoformat()))
        db.commit(); return jsonify({'ok': True, 'id': cur.lastrowid})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/gastos/generar', methods=['POST'])
def generar_gastos():
    semana_id = int(request.json['semana_id'])
    db = get_db()
    try:
        proms = ql(db.execute("SELECT * FROM promotores").fetchall())

        com_rows = ql(db.execute("""
            SELECT st.tienda_id, SUM(pt.comision) AS total
            FROM pagos_telcel pt
            JOIN subclaves_telcel st ON st.subclave = pt.subclave
            WHERE pt.semana_id = ?
            GROUP BY st.tienda_id
        """, (semana_id,)).fetchall())
        com_tienda = {r['tienda_id']: float(r['total'] or 0) for r in com_rows}

        for p in proms:
            p['tiene_seguro'] = bool(p['tiene_seguro'])
            g   = calcular_gasto_promotor(p)
            # impuestos, gastos_indirectos y fondo_contingencia son fijos por tienda;
            # se guardan en 0 por promotor y se inyectan como constante en get_gastos()
            total_promotor = round(
                g['sueldo_semanal'] + g['seguro'] + g['isn'] +
                g['aguinaldo'] + g['vacaciones'] + g['prima_vacacional'], 2
            )
            com = round(float(p.get('comisiones') or 0), 2)
            total_final = round(total_promotor + com, 2)
            db.execute("""
                INSERT INTO gastos_semana
                    (semana_id, promotor_id, sueldo_semanal, comisiones, seguro, isn,
                     impuestos, gastos_indirectos, fondo_contingencia,
                     aguinaldo, vacaciones, prima_vacacional, total)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(semana_id, promotor_id) DO UPDATE SET
                    sueldo_semanal=excluded.sueldo_semanal,
                    comisiones=excluded.comisiones,
                    seguro=excluded.seguro,
                    isn=excluded.isn,
                    impuestos=excluded.impuestos,
                    gastos_indirectos=excluded.gastos_indirectos,
                    fondo_contingencia=excluded.fondo_contingencia,
                    aguinaldo=excluded.aguinaldo,
                    vacaciones=excluded.vacaciones,
                    prima_vacacional=excluded.prima_vacacional,
                    total=excluded.total
            """, (semana_id, p['id'], g['sueldo_semanal'], com, g['seguro'],
                  g['isn'], 0, 0, 0,
                  g['aguinaldo'], g['vacaciones'], g['prima_vacacional'], total_final))

        db.commit()
        return jsonify({'ok': True, 'promotores': len(proms)})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        db.close()


@app.route('/api/gastos/<int:semana_id>')
def get_gastos(semana_id):
    db   = get_db()
    rows = ql(db.execute("""
        SELECT t.cadena, t.nombre AS tienda,
               COUNT(gs.id)             AS promotores,
               SUM(gs.sueldo_semanal)   AS sueldos,
               SUM(gs.comisiones)       AS comisiones,
               SUM(gs.seguro)           AS seguro,
               SUM(gs.aguinaldo)        AS aguinaldo,
               SUM(gs.vacaciones)       AS vacaciones,
               SUM(gs.prima_vacacional) AS prima_vacacional,
               SUM(gs.total)            AS total
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        JOIN tiendas t ON t.id = p.tienda_id
        WHERE gs.semana_id = ?
        GROUP BY t.id, t.cadena, t.nombre
        ORDER BY t.cadena, t.nombre
    """, (semana_id,)).fetchall())
    db.close()
    for r in rows:
        for k, v in r.items():
            if k not in ('cadena', 'tienda') and v is not None:
                try: r[k] = round(float(v), 2)
                except: pass
        n = int(r['promotores'])
        r['isn']                = round(61.54 * n, 2)
        r['impuestos']          = 417.05
        r['gastos_indirectos']  = 274.00
        r['fondo_contingencia'] = 27.00
        r['total']              = round(float(r['total'] or 0) + 417.05 + 274.00 + 27.00, 2)
    return jsonify(rows)


@app.route('/api/gastos/<int:semana_id>/detalle')
def get_gastos_detalle(semana_id):
    db   = get_db()
    rows = ql(db.execute("""
        SELECT gs.sueldo_semanal, gs.comisiones, gs.seguro, gs.isn,
               gs.impuestos, gs.gastos_indirectos, gs.fondo_contingencia,
               gs.aguinaldo, gs.vacaciones, gs.prima_vacacional, gs.total,
               p.nombre AS promotor, p.promotor_id,
               t.nombre AS tienda,
               t.cadena
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        JOIN tiendas t ON t.id = p.tienda_id
        WHERE gs.semana_id = ?
        ORDER BY t.cadena, t.nombre, p.promotor_id
    """, (semana_id,)).fetchall())
    db.close()
    for r in rows:
        for k, v in r.items():
            if k not in ('promotor', 'tienda', 'cadena') and v is not None:
                try: r[k] = round(float(v), 2)
                except: pass
    return jsonify(rows)


# ── TELCEL ────────────────────────────────────────────────────────────────────

@app.route('/telcel')
def telcel():
    db      = get_db()
    semanas = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return render_template('telcel.html', semanas=semanas)


@app.route('/api/telcel/upload', methods=['POST'])
def upload_telcel():
    semana_id = request.form.get('semana_id')
    if not semana_id:
        return jsonify({'ok': False, 'error': 'Selecciona una semana'}), 400
    f = request.files.get('archivo')
    if not f:
        return jsonify({'ok': False, 'error': 'No se recibió archivo'}), 400
    try:
        filename = f.filename.lower()
        raw = f.read()

        # Columnas fijas: A=Subclave, B=Concepto, C=Comisión
        all_rows = []
        if filename.endswith('.xls'):
            wb_xls = xlrd.open_workbook(file_contents=raw)
            ws_xls = wb_xls.sheet_by_index(0)
            for i in range(ws_xls.nrows):
                all_rows.append([ws_xls.cell_value(i, c) if c < ws_xls.ncols else None
                                  for c in range(3)])
        else:
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                all_rows.append(list(row[:3]))

        if not all_rows:
            return jsonify({'ok': False, 'error': 'El archivo está vacío'}), 400

        # Detectar si la primera fila es encabezado (col C no numérica)
        try:
            float(all_rows[0][2] or 0)
            start = 0
        except (TypeError, ValueError):
            start = 1

        rows_data = []
        for row in all_rows[start:]:
            if len(row) < 3:
                continue
            sub  = str(row[0]).strip() if row[0] is not None else ''
            prod = str(row[1]).strip() if row[1] is not None else ''
            try:   com = float(row[2] or 0)
            except (TypeError, ValueError): com = 0.0
            if sub and sub.lower() not in ('none', 'nan', ''):
                rows_data.append((sub, prod, com))

        if not rows_data:
            return jsonify({'ok': False, 'error': 'El archivo no contiene datos válidos'}), 400

        db = get_db()
        db.execute("DELETE FROM pagos_telcel WHERE semana_id=?", (semana_id,))
        db.executemany(
            "INSERT INTO pagos_telcel (semana_id, subclave, producto, comision) VALUES (?,?,?,?)",
            [(semana_id, r[0], r[1], r[2]) for r in rows_data])
        db.commit()

        # Conceptos únicos en orden de aparición
        conceptos = []
        seen_c = set()
        for r in rows_data:
            c = r[1]
            if c and c not in seen_c:
                seen_c.add(c)
                conceptos.append(c)

        raw_rows = ql(db.execute("""
            SELECT pt.subclave,
                   COALESCE(t.nombre, 'Sin tienda asignada') AS tienda,
                   COALESCE(t.cadena, '')                    AS cadena,
                   pt.producto,
                   SUM(pt.comision) AS subtotal
            FROM pagos_telcel pt
            LEFT JOIN subclaves_telcel st ON st.subclave = pt.subclave
            LEFT JOIN tiendas t ON t.id = st.tienda_id
            WHERE pt.semana_id = ?
            GROUP BY pt.subclave, t.nombre, t.cadena, pt.producto
            ORDER BY pt.subclave
        """, (semana_id,)).fetchall())
        db.close()

        agrupado = {}
        orden_sub = []
        for row in raw_rows:
            sub = row['subclave']
            if sub not in agrupado:
                agrupado[sub] = {'subclave': sub, 'tienda': row['tienda'],
                                 'cadena': row['cadena'], 'conceptos': {}}
                orden_sub.append(sub)
            prod = row['producto'] or ''
            agrupado[sub]['conceptos'][prod] = round(float(row['subtotal'] or 0), 2)

        resumen = []
        for sub in orden_sub:
            r = agrupado[sub]
            total = sum(r['conceptos'].values())
            resumen.append({'subclave': r['subclave'], 'tienda': r['tienda'],
                            'cadena': r['cadena'], 'conceptos': r['conceptos'],
                            'total': round(total, 2)})

        return jsonify({'ok': True, 'filas': len(rows_data), 'conceptos': conceptos, 'resumen': resumen})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── COMISIONES ────────────────────────────────────────────────────────────────

@app.route('/comisiones')
def comisiones():
    db = get_db()
    semanas = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return render_template('comisiones.html', semanas=semanas)


CONCEPTOS_FIJOS = ['AMIGO KIT', 'CBPC', 'CHIP EXPRESS', 'GARANTIZADA']

@app.route('/api/comisiones/<int:semana_id>')
def get_comisiones(semana_id):
    db = get_db()
    raw_rows = ql(db.execute("""
        SELECT COALESCE(t.cadena,'Sin cadena')           AS cadena,
               COALESCE(t.nombre,'Sin tienda asignada')  AS tienda,
               COALESCE(t.clave_principal,'')            AS clave_principal,
               pt.producto,
               SUM(pt.comision)                          AS subtotal
        FROM pagos_telcel pt
        LEFT JOIN subclaves_telcel st ON st.subclave = pt.subclave
        LEFT JOIN tiendas t ON t.id = st.tienda_id
        WHERE pt.semana_id = ?
        GROUP BY t.cadena, t.nombre, t.clave_principal, pt.producto
        ORDER BY t.cadena, t.nombre
    """, (semana_id,)).fetchall())
    db.close()

    agrupado = {}
    orden = []
    for row in raw_rows:
        key = (row['cadena'], row['tienda'])
        if key not in agrupado:
            agrupado[key] = {'cadena': row['cadena'], 'tienda': row['tienda'],
                             'clave_principal': row['clave_principal'], 'conceptos': {}}
            orden.append(key)
        prod = (row['producto'] or '').strip().upper()
        agrupado[key]['conceptos'][prod] = round(float(row['subtotal'] or 0), 2)

    rows = []
    for key in orden:
        r = agrupado[key]
        rows.append({'cadena': r['cadena'], 'tienda': r['tienda'],
                     'clave_principal': r['clave_principal'],
                     'conceptos': r['conceptos'],
                     'total': round(sum(r['conceptos'].values()), 2)})

    # Tiendas sin comisiones esa semana
    db2 = get_db()
    ids_con_rows = ql(db2.execute("""
        SELECT DISTINCT t.id FROM tiendas t
        JOIN subclaves_telcel st ON st.tienda_id = t.id
        JOIN pagos_telcel pt ON pt.subclave = st.subclave
        WHERE pt.semana_id = ?
    """, (semana_id,)).fetchall())
    ids_con = {r['id'] for r in ids_con_rows}

    todas = ql(db2.execute(
        "SELECT id, cadena, nombre, clave_principal FROM tiendas ORDER BY cadena, nombre"
    ).fetchall())
    db2.close()

    sin_comisiones = [
        {'cadena': t['cadena'], 'tienda': t['nombre'], 'clave_principal': t['clave_principal'] or ''}
        for t in todas if t['id'] not in ids_con
    ]

    return jsonify({'conceptos': CONCEPTOS_FIJOS, 'rows': rows, 'sin_comisiones': sin_comisiones})


# ── RESUMEN ───────────────────────────────────────────────────────────────────

@app.route('/resumen')
def resumen():
    db      = get_db()
    semanas = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return render_template('resumen.html', semanas=semanas)


@app.route('/api/resumen/<int:semana_id>')
def get_resumen(semana_id):
    db   = get_db()
    gast = ql(db.execute("""
        SELECT t.id, t.cadena, t.nombre AS tienda,
               SUM(gs.total) AS gastos, COUNT(gs.id) AS promotores
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        JOIN tiendas t ON t.id = p.tienda_id
        WHERE gs.semana_id = ?
        GROUP BY t.id, t.cadena, t.nombre
    """, (semana_id,)).fetchall())

    result = {}
    for r in gast:
        result[r['id']] = {
            'id': r['id'], 'cadena': r['cadena'], 'tienda': r['tienda'],
            'gastos': round(float(r['gastos'] or 0), 2),
            'promotores': int(r['promotores']), 'ingresos': 0.0
        }

    ing = ql(db.execute("""
        SELECT st.tienda_id, t.cadena, t.nombre AS tienda,
               SUM(pt.comision) AS ingresos
        FROM pagos_telcel pt
        JOIN subclaves_telcel st ON st.subclave = pt.subclave
        JOIN tiendas t ON t.id = st.tienda_id
        WHERE pt.semana_id = ?
        GROUP BY st.tienda_id, t.cadena, t.nombre
    """, (semana_id,)).fetchall())
    db.close()

    for r in ing:
        tid = r['tienda_id']
        v   = round(float(r['ingresos'] or 0), 2)
        if tid in result:
            result[tid]['ingresos'] = v
        else:
            result[tid] = {'id': tid, 'cadena': r['cadena'], 'tienda': r['tienda'],
                           'gastos': 0.0, 'promotores': 0, 'ingresos': v}

    out = sorted(result.values(), key=lambda x: (x['cadena'], x['tienda']))
    for r in out:
        r['utilidad'] = round(r['ingresos'] - r['gastos'], 2)
    return jsonify(out)


# ── BALANCE ───────────────────────────────────────────────────────────────────

@app.route('/balance')
def balance():
    db = get_db()
    semanas = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return render_template('balance.html', semanas=semanas)


@app.route('/api/balance/<int:semana_id>')
def get_balance(semana_id):
    db = get_db()

    telcel_rows = ql(db.execute("""
        SELECT UPPER(TRIM(producto)) AS producto, SUM(comision) AS total, COUNT(*) AS lineas
        FROM pagos_telcel WHERE semana_id = ?
        GROUP BY UPPER(TRIM(producto))
    """, (semana_id,)).fetchall())

    extra_rows = ql(db.execute("""
        SELECT tipo, SUM(monto) AS total
        FROM comisiones_extra WHERE semana_id = ?
        GROUP BY tipo
    """, (semana_id,)).fetchall())

    g = q1(db.execute("""
        SELECT COUNT(gs.id)                                       AS n_prom,
               COUNT(DISTINCT p.tienda_id)                        AS n_tiend,
               SUM(gs.sueldo_semanal)                             AS sueldos,
               SUM(gs.comisiones)                                 AS comisiones,
               SUM(CASE WHEN gs.comisiones > 0 THEN 1 ELSE 0 END) AS n_com,
               SUM(gs.seguro)                                     AS seguro,
               SUM(CASE WHEN gs.seguro > 0 THEN 1 ELSE 0 END)    AS n_seg,
               SUM(gs.aguinaldo)                                  AS aguinaldo,
               SUM(gs.vacaciones)                                 AS vacaciones,
               SUM(CASE WHEN gs.vacaciones > 0 THEN 1 ELSE 0 END) AS n_vac,
               SUM(gs.prima_vacacional)                           AS prima_vacacional,
               SUM(CASE WHEN gs.prima_vacacional > 0 THEN 1 ELSE 0 END) AS n_pv
        FROM gastos_semana gs
        JOIN promotores p ON p.id = gs.promotor_id
        WHERE gs.semana_id = ?
    """, (semana_id,)).fetchone())
    db.close()

    if not g or not g['n_prom']:
        return jsonify({'ok': False, 'error': 'Sin gastos generados para esta semana'})

    np  = int(g['n_prom']  or 0)
    nt  = int(g['n_tiend'] or 0)

    telcel = {}
    telcel_lineas = {}
    for r in telcel_rows:
        prod = (r['producto'] or '').strip()
        if prod:
            telcel[prod] = round(float(r['total'] or 0), 2)
            telcel_lineas[prod] = int(r['lineas'] or 0)

    extra = {r['tipo']: round(float(r['total'] or 0), 2) for r in extra_rows}

    total_ingresos = round(sum(telcel.values()) + sum(extra.values()), 2)

    sueldos = round(float(g['sueldos'] or 0), 2)
    coms    = round(float(g['comisiones'] or 0), 2)
    seguro  = round(float(g['seguro'] or 0), 2)
    isn     = round(61.54  * np, 2)
    imp     = round(417.05 * nt, 2)
    gi      = round(274.00 * nt, 2)
    fc      = round(27.00  * np, 2)
    ag      = round(float(g['aguinaldo'] or 0), 2)
    vac     = round(float(g['vacaciones'] or 0), 2)
    pv      = round(float(g['prima_vacacional'] or 0), 2)

    total_gastos   = round(sueldos + coms + seguro + isn + imp + gi + fc + ag + vac + pv, 2)
    utilidad       = round(total_ingresos - total_gastos, 2)
    margen         = round(utilidad / total_ingresos * 100, 1) if total_ingresos else 0.0

    return jsonify({
        'ok': True,
        'ingresos': {'telcel': telcel, 'telcel_lineas': telcel_lineas, 'extra': extra, 'total': total_ingresos},
        'gastos': {
            'sueldos': sueldos,   'n_prom': np,
            'comisiones': coms,   'n_com': int(g['n_com'] or 0),
            'seguro': seguro,     'n_seg': int(g['n_seg'] or 0),
            'isn': isn,
            'impuestos': imp,     'n_tiend': nt,
            'gastos_indirectos': gi,
            'fondo_contingencia': fc,
            'aguinaldo': ag,
            'vacaciones': vac,    'n_vac': int(g['n_vac'] or 0),
            'prima_vacacional': pv,'n_pv': int(g['n_pv'] or 0),
            'total': total_gastos,
        },
        'utilidad': utilidad,
        'margen': margen,
    })


# ── COMISIONES EXTRA ─────────────────────────────────────────────────────────

TIPOS_COMISION_EXTRA = ['2% AP', 'Volumen Garantizado', 'Amigo Kit', 'Cadena Comercial']

@app.route('/comisiones-extra')
def comisiones_extra():
    db = get_db()
    semanas = ql(db.execute("SELECT * FROM semanas ORDER BY fecha_inicio DESC").fetchall())
    db.close()
    return render_template('comisiones_extra.html', semanas=semanas,
                           tipos=TIPOS_COMISION_EXTRA)


@app.route('/api/comisiones-extra/<int:semana_id>')
def get_comisiones_extra(semana_id):
    db = get_db()
    rows = ql(db.execute("""
        SELECT id, tipo, monto, notas
        FROM comisiones_extra
        WHERE semana_id = ?
        ORDER BY tipo
    """, (semana_id,)).fetchall())
    db.close()
    for r in rows:
        r['monto'] = round(float(r['monto'] or 0), 2)
    return jsonify(rows)


@app.route('/api/comisiones-extra', methods=['POST'])
def add_comision_extra():
    d = request.json
    db = get_db()
    try:
        cur = db.execute("""
            INSERT INTO comisiones_extra (semana_id, tipo, monto, notas)
            VALUES (?,?,?,?)
        """, (int(d['semana_id']), d['tipo'].strip(),
              float(d['monto']), (d.get('notas') or '').strip() or None))
        db.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/comisiones-extra/<int:cid>', methods=['PUT'])
def edit_comision_extra(cid):
    d = request.json
    db = get_db()
    try:
        db.execute("""
            UPDATE comisiones_extra SET tipo=?, monto=?, notas=?
            WHERE id=?
        """, (d['tipo'].strip(), float(d['monto']),
              (d.get('notas') or '').strip() or None, cid))
        db.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/api/comisiones-extra/<int:cid>', methods=['DELETE'])
def del_comision_extra(cid):
    db = get_db()
    try:
        db.execute("DELETE FROM comisiones_extra WHERE id=?", (cid,))
        db.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback(); return jsonify({'ok': False, 'error': str(e)}), 400
    finally:
        db.close()


@app.route('/')
def index():
    return redirect('/gastos')


# ── Arranque ──────────────────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    print()
    print("=" * 52)
    print("  gastos-cadenas  —  modo LOCAL (SQLite)")
    print("  Abre en tu navegador: http://127.0.0.1:5000")
    print("=" * 52)
    print()
    app.run(debug=True, host='127.0.0.1', port=5000, use_reloader=False)
